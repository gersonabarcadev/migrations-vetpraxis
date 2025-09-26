#!/usr/bin/env python3
"""
Generador de templates Excel para importación de vacunas
Basado en análisis de pestañas 'vacunas' y 'pacientevacuna'

Formato de salida:
A => clinic_record_import_id (generado)
B => PatientId 
C => DataDate (fecha de aplicación)
D => Razón de atención: "Vacuna"
E => Tratamiento: nombre de la vacuna
F => Cantidad: 1 (fijo)
G => Notas: notas adicionales
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_clinic_record_id(patient_id, date_str):
    """
    Generar ID consistente para historia clínica basado en paciente y fecha
    Mantiene consistencia con sistema de apuntes y procedimientos
    """
    # Crear string único para el hash
    unique_string = f"VACUNA_{patient_id}_{date_str}"
    
    # Generar hash MD5 y tomar primeros 8 caracteres
    hash_object = hashlib.md5(unique_string.encode())
    hash_hex = hash_object.hexdigest()[:8]
    
    # Convertir a número entero positivo
    clinic_record_id = int(hash_hex, 16) % 100000000  # Limitar a 8 dígitos
    
    return clinic_record_id

def load_and_prepare_data():
    """Cargar y preparar datos de vacunas"""
    print("📖 Cargando datos de vacunas...")
    
    new_file = "/Users/enrique/Proyectos/imports/source/cuvet-v2.xlsx"
    
    if not os.path.exists(new_file):
        raise FileNotFoundError(f"Archivo no encontrado: {new_file}")
    
    # Cargar ambas pestañas
    df_vacunas = pd.read_excel(new_file, sheet_name='vacunas', engine='openpyxl')
    df_pacientevacuna = pd.read_excel(new_file, sheet_name='pacientevacuna', engine='openpyxl')
    
    print(f"✅ Catálogo de vacunas: {len(df_vacunas):,} registros")
    print(f"✅ Aplicaciones de vacunas: {len(df_pacientevacuna):,} registros")
    
    # Filtrar solo registros activos (no eliminados)
    if 'IsDeleted' in df_pacientevacuna.columns:
        df_active = df_pacientevacuna[df_pacientevacuna['IsDeleted'] == 0].copy()
        eliminados = len(df_pacientevacuna) - len(df_active)
        print(f"🗑️  Eliminados: {eliminados:,} registros ({eliminados/len(df_pacientevacuna)*100:.1f}%)")
    else:
        df_active = df_pacientevacuna.copy()
    
    print(f"📊 Registros activos para procesar: {len(df_active):,}")
    
    # Convertir fechas
    df_active['DataDate'] = pd.to_datetime(df_active['DataDate'])
    
    # Crear diccionario de nombres de vacunas
    vaccine_names = dict(zip(df_vacunas['VaccineId'], df_vacunas['Name']))
    
    print(f"🔗 Catálogo de vacunas disponible: {len(vaccine_names)} tipos")
    
    return df_active, vaccine_names

def process_vaccine_records(df_active, vaccine_names):
    """Procesar registros de vacunas para importación"""
    print("\n🔄 Procesando registros de vacunas...")
    
    records = []
    
    for _, row in df_active.iterrows():
        try:
            # Datos básicos
            patient_id = int(row['PatientId'])
            data_date = row['DataDate']
            vaccine_id = int(row['VaccineId'])
            note = row.get('Note', '') if pd.notna(row.get('Note', '')) else ''
            
            # Formatear fecha para ID consistente
            date_str = data_date.strftime('%Y%m%d')
            
            # Generar clinic_record_import_id
            clinic_record_id = generate_clinic_record_id(patient_id, date_str)
            
            # Obtener nombre de vacuna
            vaccine_name = vaccine_names.get(vaccine_id, f"Vacuna ID {vaccine_id}")
            
            # Crear registro
            record = {
                'clinic_record_import_id': clinic_record_id,
                'PatientId': patient_id,
                'DataDate': data_date,
                'Razon': 'Vacuna',
                'Tratamiento': vaccine_name,
                'Cantidad': 1,
                'Notas': note
            }
            
            records.append(record)
            
        except Exception as e:
            print(f"⚠️  Error procesando registro: {e}")
            continue
    
    df_result = pd.DataFrame(records)
    print(f"✅ Procesados exitosamente: {len(df_result):,} registros")
    
    return df_result

def create_excel_template(df_data, output_dir):
    """Crear archivo Excel con formato específico"""
    print(f"\n📝 Generando template Excel...")
    
    # Crear directorio si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Ordenar por fecha y paciente
    df_sorted = df_data.sort_values(['DataDate', 'PatientId'])
    
    # Configuración de archivos
    records_per_file = 10000  # 10K registros por archivo como en apuntes
    total_records = len(df_sorted)
    num_files = (total_records + records_per_file - 1) // records_per_file
    
    print(f"📊 Total registros: {total_records:,}")
    print(f"📁 Archivos a generar: {num_files}")
    print(f"📄 Registros por archivo: {records_per_file:,}")
    
    generated_files = []
    
    for file_num in range(num_files):
        start_idx = file_num * records_per_file
        end_idx = min(start_idx + records_per_file, total_records)
        df_chunk = df_sorted.iloc[start_idx:end_idx]
        
        # Nombre del archivo
        filename = f"vacunas_import_{file_num + 1:02d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Vacunas Import"
        
        # Headers
        headers = [
            'A: ID Historia Clínica',
            'B: ID Mascota', 
            'C: Fecha de Atención',
            'D: Razón de Atención',
            'E: Tratamiento',
            'F: Cantidad',
            'G: Notas'
        ]
        
        # Estilo para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row_idx, (_, record) in enumerate(df_chunk.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=record['clinic_record_import_id'])
            ws.cell(row=row_idx, column=2, value=record['PatientId'])
            ws.cell(row=row_idx, column=3, value=record['DataDate'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=4, value=record['Razon'])
            ws.cell(row=row_idx, column=5, value=record['Tratamiento'])
            ws.cell(row=row_idx, column=6, value=record['Cantidad'])
            ws.cell(row=row_idx, column=7, value=record['Notas'])
        
        # Ajustar ancho de columnas
        column_widths = [20, 15, 20, 18, 40, 10, 50]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Guardar archivo
        wb.save(filepath)
        generated_files.append(filepath)
        
        print(f"   ✅ {filename}: {len(df_chunk):,} registros")
    
    return generated_files

def generate_summary_report(df_data, output_dir):
    """Generar reporte resumen"""
    print(f"\n📊 Generando reporte resumen...")
    
    summary_file = os.path.join(output_dir, "resumen_vacunas_import.txt")
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("RESUMEN DE IMPORTACIÓN - VACUNAS\n")
        f.write("=" * 50 + "\n\n")
        
        f.write(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total de registros procesados: {len(df_data):,}\n\n")
        
        # Estadísticas por fecha
        f.write("DISTRIBUCIÓN POR AÑO:\n")
        year_counts = df_data['DataDate'].dt.year.value_counts().sort_index()
        for year, count in year_counts.items():
            f.write(f"  {year}: {count:,} aplicaciones\n")
        
        f.write(f"\nRANGO DE FECHAS:\n")
        f.write(f"  Desde: {df_data['DataDate'].min()}\n")
        f.write(f"  Hasta: {df_data['DataDate'].max()}\n")
        
        # Top vacunas
        f.write(f"\nTOP 10 VACUNAS MÁS APLICADAS:\n")
        top_vaccines = df_data['Tratamiento'].value_counts().head(10)
        for vaccine, count in top_vaccines.items():
            f.write(f"  {vaccine}: {count:,} aplicaciones\n")
        
        # Estadísticas de pacientes
        f.write(f"\nESTADÍSTICAS DE PACIENTES:\n")
        f.write(f"  Pacientes únicos: {df_data['PatientId'].nunique():,}\n")
        f.write(f"  Promedio vacunas por paciente: {len(df_data)/df_data['PatientId'].nunique():.2f}\n")
        
        # Notas
        notes_count = df_data['Notas'].astype(str).str.strip().ne('').sum()
        f.write(f"\nNOTAS:\n")
        f.write(f"  Registros con notas: {notes_count:,} ({notes_count/len(df_data)*100:.1f}%)\n")
        
        f.write(f"\nFORMATO DE ARCHIVOS:\n")
        f.write(f"  Columna A: ID Historia Clínica (generado automáticamente)\n")
        f.write(f"  Columna B: ID Mascota (PatientId original)\n")
        f.write(f"  Columna C: Fecha de Atención\n")
        f.write(f"  Columna D: Razón de Atención (siempre 'Vacuna')\n")
        f.write(f"  Columna E: Tratamiento (nombre de la vacuna)\n")
        f.write(f"  Columna F: Cantidad (siempre 1)\n")
        f.write(f"  Columna G: Notas adicionales\n")
    
    print(f"✅ Reporte guardado: {summary_file}")

def main():
    print("🩹 GENERADOR DE TEMPLATES - VACUNAS")
    print("=" * 50)
    
    output_dir = "/Users/enrique/Proyectos/imports/generated_files/vacunas"
    
    try:
        # 1. Cargar datos
        df_active, vaccine_names = load_and_prepare_data()
        
        # 2. Procesar registros
        df_result = process_vaccine_records(df_active, vaccine_names)
        
        if len(df_result) == 0:
            print("❌ No hay registros para procesar")
            return
        
        # 3. Crear templates Excel
        generated_files = create_excel_template(df_result, output_dir)
        
        # 4. Generar reporte resumen
        generate_summary_report(df_result, output_dir)
        
        # 5. Resumen final
        print(f"\n🎉 ¡GENERACIÓN COMPLETADA!")
        print(f"📁 Directorio: {output_dir}")
        print(f"📄 Archivos generados: {len(generated_files)}")
        print(f"📊 Total registros: {len(df_result):,}")
        print(f"🏥 Pacientes únicos: {df_result['PatientId'].nunique():,}")
        print(f"💉 Tipos de vacunas: {df_result['Tratamiento'].nunique()}")
        
        # Mostrar archivos generados
        print(f"\n📋 Archivos creados:")
        for filepath in generated_files:
            filename = os.path.basename(filepath)
            print(f"   • {filename}")
        print(f"   • resumen_vacunas_import.txt")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
