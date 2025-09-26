#!/usr/bin/env python3
"""
Generador de templates Excel para importación de diagnósticos
Nueva estructura específica con IDs correlativos

Formato de salida:
A => import_clinic_record_id (correlativo desde 99869913)
B => PatientId 
C => DataDate (fecha de diagnóstico)
D => Note (diagnosticos.Name + [PARRAFO] + pacientediagnosticos.Note)
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_and_prepare_data():
    """Cargar y preparar datos de diagnósticos"""
    print("📖 Cargando datos de diagnósticos...")
    
    new_file = "/Users/enrique/Proyectos/imports/source/cuvet-v2.xlsx"
    
    if not os.path.exists(new_file):
        raise FileNotFoundError(f"Archivo no encontrado: {new_file}")
    
    # Cargar ambas pestañas
    df_diagnosticos = pd.read_excel(new_file, sheet_name='diagnosticos', engine='openpyxl')
    df_pacientediagnosticos = pd.read_excel(new_file, sheet_name='pacientediagnosticos', engine='openpyxl')
    
    print(f"✅ Catálogo de diagnósticos: {len(df_diagnosticos):,} registros")
    print(f"✅ Aplicaciones de diagnósticos: {len(df_pacientediagnosticos):,} registros")
    
    # Filtrar solo registros activos (no eliminados)
    if 'IsDeleted' in df_pacientediagnosticos.columns:
        df_active = df_pacientediagnosticos[df_pacientediagnosticos['IsDeleted'] == 0].copy()
        eliminados = len(df_pacientediagnosticos) - len(df_active)
        print(f"🗑️  Eliminados: {eliminados:,} registros ({eliminados/len(df_pacientediagnosticos)*100:.1f}%)")
    else:
        df_active = df_pacientediagnosticos.copy()
    
    print(f"📊 Registros activos para procesar: {len(df_active):,}")
    
    # Convertir fechas
    df_active['DataDate'] = pd.to_datetime(df_active['DataDate'])
    
    # Crear diccionario de nombres de diagnósticos
    diagnostic_names = dict(zip(df_diagnosticos['DiagnosticId'], df_diagnosticos['Name']))
    
    print(f"🔗 Catálogo de diagnósticos disponible: {len(diagnostic_names)} tipos")
    
    return df_active, diagnostic_names

def process_diagnostic_records(df_active, diagnostic_names):
    """Procesar registros de diagnósticos para importación"""
    print("\n🔄 Procesando registros de diagnósticos...")
    
    records = []
    current_id = 99869913  # ID inicial correlativo (siguiente después de procedimientos)
    
    # Ordenar por fecha para mantener secuencia lógica
    df_sorted = df_active.sort_values(['DataDate', 'PatientId']).reset_index(drop=True)
    
    for idx, row in df_sorted.iterrows():
        try:
            # Datos básicos
            patient_id = int(row['PatientId'])
            data_date = row['DataDate']
            diagnostic_id = int(row['DiagnosticId'])
            note = row.get('Note', '') if pd.notna(row.get('Note', '')) else ''
            
            # Obtener nombre del diagnóstico
            diagnostic_name = diagnostic_names.get(diagnostic_id, f"Diagnóstico ID {diagnostic_id}")
            
            # Construir campo Note según especificación
            if note.strip():
                # Hay nota: Nombre + [PARRAFO] + Nota
                formatted_note = f"{diagnostic_name}[PARRAFO]{note.strip()}"
            else:
                # Sin nota: Solo nombre
                formatted_note = diagnostic_name
            
            # Crear registro
            record = {
                'import_clinic_record_id': current_id,
                'PatientId': patient_id,
                'DataDate': data_date,
                'Note': formatted_note
            }
            
            records.append(record)
            current_id += 1
            
            # Progreso cada 5000 registros
            if (idx + 1) % 5000 == 0:
                print(f"   Procesados: {idx + 1:,}/{len(df_sorted):,} registros...")
            
        except Exception as e:
            print(f"⚠️  Error procesando registro {idx}: {e}")
            continue
    
    df_result = pd.DataFrame(records)
    print(f"✅ Procesados exitosamente: {len(df_result):,} registros")
    print(f"🆔 IDs asignados: {99869913} - {current_id - 1}")
    
    return df_result

def create_excel_template(df_data, output_dir):
    """Crear archivo Excel con formato específico"""
    print(f"\n📝 Generando template Excel...")
    
    # Crear directorio si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Configuración de archivos
    records_per_file = 10000  # 10K registros por archivo
    total_records = len(df_data)
    num_files = (total_records + records_per_file - 1) // records_per_file
    
    print(f"📊 Total registros: {total_records:,}")
    print(f"📁 Archivos a generar: {num_files}")
    print(f"📄 Registros por archivo: {records_per_file:,}")
    
    generated_files = []
    
    for file_num in range(num_files):
        start_idx = file_num * records_per_file
        end_idx = min(start_idx + records_per_file, total_records)
        df_chunk = df_data.iloc[start_idx:end_idx]
        
        # Nombre del archivo
        filename = f"diagnosticos_import_{file_num + 1:02d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Diagnosticos Import"
        
        # Headers
        headers = [
            'A: Import Clinic Record ID',
            'B: Import Pet ID',
            'C: Import Date',
            'D: Note'
        ]
        
        # Estilo para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row_idx, (_, record) in enumerate(df_chunk.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=record['import_clinic_record_id'])
            ws.cell(row=row_idx, column=2, value=record['PatientId'])
            ws.cell(row=row_idx, column=3, value=record['DataDate'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=4, value=record['Note'])
        
        # Ajustar ancho de columnas
        column_widths = [25, 15, 20, 80]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Guardar archivo
        wb.save(filepath)
        generated_files.append(filepath)
        
        print(f"   ✅ {filename}: {len(df_chunk):,} registros (IDs: {df_chunk.iloc[0]['import_clinic_record_id']} - {df_chunk.iloc[-1]['import_clinic_record_id']})")
    
    return generated_files

def generate_summary_report(df_data, output_dir):
    """Generar reporte resumen"""
    print(f"\n📊 Generando reporte resumen...")
    
    summary_file = os.path.join(output_dir, "resumen_diagnosticos_import.txt")
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("RESUMEN DE IMPORTACIÓN - DIAGNÓSTICOS\n")
        f.write("=" * 50 + "\n\n")
        
        f.write(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total de registros procesados: {len(df_data):,}\n")
        f.write(f"Rango de IDs asignados: {df_data['import_clinic_record_id'].min()} - {df_data['import_clinic_record_id'].max()}\n\n")
        
        # Estadísticas por fecha
        f.write("DISTRIBUCIÓN POR AÑO:\n")
        year_counts = df_data['DataDate'].dt.year.value_counts().sort_index()
        for year, count in year_counts.items():
            f.write(f"  {year}: {count:,} diagnósticos\n")
        
        f.write(f"\nRANGO DE FECHAS:\n")
        f.write(f"  Desde: {df_data['DataDate'].min()}\n")
        f.write(f"  Hasta: {df_data['DataDate'].max()}\n")
        
        # Análisis de notas con/sin contenido adicional
        notes_with_paragraph = df_data['Note'].str.contains(r'\[PARRAFO\]', na=False).sum()
        notes_only_name = len(df_data) - notes_with_paragraph
        
        f.write(f"\nESTADÍSTICAS DE NOTAS:\n")
        f.write(f"  Solo nombre del diagnóstico: {notes_only_name:,} ({notes_only_name/len(df_data)*100:.1f}%)\n")
        f.write(f"  Con notas adicionales: {notes_with_paragraph:,} ({notes_with_paragraph/len(df_data)*100:.1f}%)\n")
        
        # Top diagnósticos
        # Extraer nombres de diagnósticos de las notas
        diagnostic_names = []
        for note in df_data['Note']:
            if '[PARRAFO]' in note:
                diagnostic_name = note.split('[PARRAFO]')[0]
            else:
                diagnostic_name = note
            diagnostic_names.append(diagnostic_name)
        
        f.write(f"\nTOP 10 DIAGNÓSTICOS MÁS FRECUENTES:\n")
        diagnostic_counts = pd.Series(diagnostic_names).value_counts().head(10)
        for diagnostic, count in diagnostic_counts.items():
            f.write(f"  {diagnostic}: {count:,} diagnósticos\n")
        
        # Estadísticas de pacientes
        f.write(f"\nESTADÍSTICAS DE PACIENTES:\n")
        f.write(f"  Pacientes únicos: {df_data['PatientId'].nunique():,}\n")
        f.write(f"  Promedio diagnósticos por paciente: {len(df_data)/df_data['PatientId'].nunique():.2f}\n")
        
        f.write(f"\nFORMATO DE ARCHIVOS:\n")
        f.write(f"  Columna A: Import Clinic Record ID (correlativo desde 99869913)\n")
        f.write(f"  Columna B: Import Pet ID (PatientId original)\n")
        f.write(f"  Columna C: Import Date (fecha de diagnóstico)\n")
        f.write(f"  Columna D: Note (Nombre[PARRAFO]Nota o solo Nombre)\n")
        f.write(f"\nSEPARADOR:\n")
        f.write(f"  [PARRAFO] - Para separar nombre del diagnóstico de notas adicionales\n")
        f.write(f"  Se puede reemplazar con saltos de línea en base de datos\n")
    
    print(f"✅ Reporte guardado: {summary_file}")

def generate_validation_samples(df_data, output_dir):
    """Generar muestras para validación"""
    print(f"\n🔍 Generando muestras de validación...")
    
    samples_file = os.path.join(output_dir, "muestras_validacion.txt")
    
    with open(samples_file, 'w', encoding='utf-8') as f:
        f.write("MUESTRAS DE VALIDACIÓN - DIAGNÓSTICOS\n")
        f.write("=" * 50 + "\n\n")
        
        # Muestras con notas adicionales
        with_notes = df_data[df_data['Note'].str.contains(r'\[PARRAFO\]', na=False)].head(5)
        f.write("EJEMPLOS CON NOTAS ADICIONALES:\n")
        f.write("-" * 30 + "\n")
        for i, (_, row) in enumerate(with_notes.iterrows(), 1):
            f.write(f"Ejemplo {i}:\n")
            f.write(f"  ID: {row['import_clinic_record_id']}\n")
            f.write(f"  Paciente: {row['PatientId']}\n")
            f.write(f"  Fecha: {row['DataDate']}\n")
            f.write(f"  Nota completa: {row['Note'][:200]}{'...' if len(row['Note']) > 200 else ''}\n\n")
        
        # Muestras solo con nombre
        without_notes = df_data[~df_data['Note'].str.contains(r'\[PARRAFO\]', na=False)].head(5)
        f.write("EJEMPLOS SOLO CON NOMBRE DE DIAGNÓSTICO:\n")
        f.write("-" * 40 + "\n")
        for i, (_, row) in enumerate(without_notes.iterrows(), 1):
            f.write(f"Ejemplo {i}:\n")
            f.write(f"  ID: {row['import_clinic_record_id']}\n")
            f.write(f"  Paciente: {row['PatientId']}\n")
            f.write(f"  Fecha: {row['DataDate']}\n")
            f.write(f"  Nota: {row['Note']}\n\n")
    
    print(f"✅ Muestras guardadas: {samples_file}")

def main():
    print("🏥 GENERADOR DE TEMPLATES - DIAGNÓSTICOS")
    print("=" * 50)
    
    output_dir = "/Users/enrique/Proyectos/imports/generated_files/diagnosticos"
    
    try:
        # 1. Cargar datos
        df_active, diagnostic_names = load_and_prepare_data()
        
        # 2. Procesar registros
        df_result = process_diagnostic_records(df_active, diagnostic_names)
        
        if len(df_result) == 0:
            print("❌ No hay registros para procesar")
            return
        
        # 3. Crear templates Excel
        generated_files = create_excel_template(df_result, output_dir)
        
        # 4. Generar reporte resumen
        generate_summary_report(df_result, output_dir)
        
        # 5. Generar muestras de validación
        generate_validation_samples(df_result, output_dir)
        
        # 6. Resumen final
        print(f"\n🎉 ¡GENERACIÓN COMPLETADA!")
        print(f"📁 Directorio: {output_dir}")
        print(f"📄 Archivos generados: {len(generated_files)}")
        print(f"📊 Total registros: {len(df_result):,}")
        print(f"🏥 Pacientes únicos: {df_result['PatientId'].nunique():,}")
        print(f"🆔 Rango IDs: {df_result['import_clinic_record_id'].min()} - {df_result['import_clinic_record_id'].max()}")
        
        # Estadísticas de formato
        with_notes = df_result['Note'].str.contains(r'\[PARRAFO\]', na=False).sum()
        without_notes = len(df_result) - with_notes
        print(f"📝 Solo nombre: {without_notes:,} ({without_notes/len(df_result)*100:.1f}%)")
        print(f"📝 Con notas adicionales: {with_notes:,} ({with_notes/len(df_result)*100:.1f}%)")
        
        # Mostrar archivos generados
        print(f"\n📋 Archivos creados:")
        for filepath in generated_files:
            filename = os.path.basename(filepath)
            print(f"   • {filename}")
        print(f"   • resumen_diagnosticos_import.txt")
        print(f"   • muestras_validacion.txt")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
