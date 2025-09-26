#!/usr/bin/env python3
"""
Generador de templates Excel para importación de prescripciones
Agrupa por PatientId + DataDate y concatena múltiples medicamentos

Formato de salida:
A => import_clinic_record_id (correlativo desde 99883852)
B => PatientId 
C => DataDate (fecha de prescripción)
D => Note (Name + Description + [PARRAFO] + RequestedUsage + AmountToBuy, concatenado para misma fecha)
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def load_and_prepare_data():
    """Cargar y preparar datos de prescripciones"""
    print("📖 Cargando datos de prescripciones...")
    
    source_file = "/Users/enrique/Proyectos/imports/source/cuvet-v2.xlsx"
    
    if not os.path.exists(source_file):
        raise FileNotFoundError(f"Archivo no encontrado: {source_file}")
    
    # Cargar datos de prescripcion
    df_prescripcion = pd.read_excel(source_file, sheet_name='prescripcion', engine='openpyxl')
    
    print(f"✅ Datos de prescripciones: {len(df_prescripcion):,} registros")
    
    # Filtrar solo registros activos (no eliminados)
    if 'IsDeleted' in df_prescripcion.columns:
        df_active = df_prescripcion[df_prescripcion['IsDeleted'] == 0].copy()
        eliminados = len(df_prescripcion) - len(df_active)
        print(f"🗑️  Eliminados: {eliminados:,} registros ({eliminados/len(df_prescripcion)*100:.1f}%)")
    else:
        df_active = df_prescripcion.copy()
    
    print(f"📊 Registros activos para procesar: {len(df_active):,}")
    
    # Convertir fechas
    df_active['DataDate'] = pd.to_datetime(df_active['DataDate'])
    
    # Limpiar y preparar campos
    df_active['Name'] = df_active['Name'].fillna('').astype(str)
    df_active['Description'] = df_active['Description'].fillna('').astype(str)
    df_active['RequestedUsage'] = df_active['RequestedUsage'].fillna('').astype(str)
    df_active['AmountToBuy'] = df_active['AmountToBuy'].fillna('').astype(str)
    
    return df_active

def group_prescriptions_by_datetime(df_active):
    """Agrupar prescripciones por PatientId + DataDate"""
    print("\n🔄 Agrupando prescripciones por paciente y fecha...")
    
    # Agrupar por PatientId y DataDate
    grouped = df_active.groupby(['PatientId', 'DataDate'])
    
    print(f"📊 Total grupos (PatientId + DataDate): {len(grouped):,}")
    
    # Estadísticas de agrupación
    group_sizes = grouped.size()
    print(f"📈 Estadísticas de agrupación:")
    print(f"   Grupos con 1 medicamento: {(group_sizes == 1).sum():,}")
    print(f"   Grupos con 2-5 medicamentos: {((group_sizes >= 2) & (group_sizes <= 5)).sum():,}")
    print(f"   Grupos con 6-10 medicamentos: {((group_sizes >= 6) & (group_sizes <= 10)).sum():,}")
    print(f"   Grupos con >10 medicamentos: {(group_sizes > 10).sum():,}")
    print(f"   Máximo medicamentos en un grupo: {group_sizes.max()}")
    
    return grouped

def create_prescription_note(medications_group):
    """Crear nota de prescripción con formato específico"""
    medication_parts = []
    
    for _, row in medications_group.iterrows():
        # Componentes del medicamento
        name = row['Name'].strip()
        description = row['Description'].strip()
        requested_usage = row['RequestedUsage'].strip()
        amount_to_buy = row['AmountToBuy'].strip()
        
        # Construir parte del medicamento
        med_part = name
        
        if description:
            med_part += f"\n{description}"
        
        # Construir dosificación/cantidad
        dosage_parts = []
        if requested_usage:
            dosage_parts.append(requested_usage)
        if amount_to_buy:
            dosage_parts.append(amount_to_buy)
        
        if dosage_parts:
            dosage = " + ".join(dosage_parts)
            med_part += f"\n{dosage}"
        
        medication_parts.append(med_part)
    
    # Unir todas las partes con [PARRAFO]
    complete_note = "[PARRAFO]".join(medication_parts)
    
    return complete_note

def process_prescription_records(grouped_data):
    """Procesar registros agrupados de prescripciones para importación"""
    print("\n🔄 Procesando registros agrupados de prescripciones...")
    
    records = []
    current_id = 99883852  # ID inicial correlativo
    
    # Procesar cada grupo
    group_count = 0
    total_groups = len(grouped_data)
    
    for (patient_id, data_date), group in grouped_data:
        try:
            # Crear nota de prescripción combinada
            combined_note = create_prescription_note(group)
            
            # Crear registro
            record = {
                'import_clinic_record_id': current_id,
                'PatientId': int(patient_id),
                'DataDate': data_date,
                'Note': combined_note,
                'medications_count': len(group)
            }
            
            records.append(record)
            current_id += 1
            group_count += 1
            
            # Progreso cada 1000 grupos
            if group_count % 1000 == 0:
                print(f"   Procesados: {group_count:,}/{total_groups:,} grupos...")
            
        except Exception as e:
            print(f"⚠️  Error procesando grupo {patient_id}-{data_date}: {e}")
            continue
    
    df_result = pd.DataFrame(records)
    print(f"✅ Procesados exitosamente: {len(df_result):,} registros agrupados")
    print(f"🆔 IDs asignados: {99883852} - {current_id - 1}")
    
    # Estadísticas de medicamentos por registro
    med_stats = df_result['medications_count'].describe()
    print(f"📊 Medicamentos por registro:")
    print(f"   Promedio: {med_stats['mean']:.2f}")
    print(f"   Mediana: {med_stats['50%']:.0f}")
    print(f"   Máximo: {med_stats['max']:.0f}")
    
    return df_result

def create_excel_template(df_data, output_dir):
    """Crear archivo Excel con formato específico"""
    print(f"\n📝 Generando template Excel...")
    
    # Crear directorio si no existe
    os.makedirs(output_dir, exist_ok=True)
    
    # Configuración de archivos
    records_per_file = 10000  # 10K registros por archivo
    total_records = len(df_data)
    num_files = (total_records + records_per_file - 1) // records_per_file
    
    print(f"📊 Total registros: {total_records:,}")
    print(f"📁 Archivos a generar: {num_files}")
    print(f"📄 Registros por archivo: {records_per_file:,}")
    
    generated_files = []
    
    for file_num in range(num_files):
        start_idx = file_num * records_per_file
        end_idx = min(start_idx + records_per_file, total_records)
        df_chunk = df_data.iloc[start_idx:end_idx]
        
        # Nombre del archivo
        filename = f"prescripciones_import_{file_num + 1:02d}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Prescripciones Import"
        
        # Headers
        headers = [
            'A: Import Clinic Record ID',
            'B: Import Pet ID',
            'C: Import Date',
            'D: Note'
        ]
        
        # Estilo para headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Escribir datos
        for row_idx, (_, record) in enumerate(df_chunk.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=record['import_clinic_record_id'])
            ws.cell(row=row_idx, column=2, value=record['PatientId'])
            ws.cell(row=row_idx, column=3, value=record['DataDate'].strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_idx, column=4, value=record['Note'])
        
        # Ajustar ancho de columnas
        column_widths = [25, 15, 20, 100]  # Columna D más ancha para prescripciones
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Guardar archivo
        wb.save(filepath)
        generated_files.append(filepath)
        
        print(f"   ✅ {filename}: {len(df_chunk):,} registros (IDs: {df_chunk.iloc[0]['import_clinic_record_id']} - {df_chunk.iloc[-1]['import_clinic_record_id']})")
    
    return generated_files

def generate_summary_report(df_data, df_original, output_dir):
    """Generar reporte resumen"""
    print(f"\n📊 Generando reporte resumen...")
    
    summary_file = os.path.join(output_dir, "resumen_prescripciones_import.txt")
    
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write("RESUMEN DE IMPORTACIÓN - PRESCRIPCIONES\n")
        f.write("=" * 50 + "\n\n")
        
        f.write(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"Total de registros agrupados: {len(df_data):,}\n")
        f.write(f"Total de prescripciones originales: {len(df_original):,}\n")
        f.write(f"Factor de compresión: {len(df_original)/len(df_data):.2f}x\n")
        f.write(f"Rango de IDs asignados: {df_data['import_clinic_record_id'].min()} - {df_data['import_clinic_record_id'].max()}\n\n")
        
        # Estadísticas por fecha
        f.write("DISTRIBUCIÓN POR AÑO:\n")
        year_counts = df_data['DataDate'].dt.year.value_counts().sort_index()
        for year, count in year_counts.items():
            f.write(f"  {year}: {count:,} prescripciones agrupadas\n")
        
        f.write(f"\nRANGO DE FECHAS:\n")
        f.write(f"  Desde: {df_data['DataDate'].min()}\n")
        f.write(f"  Hasta: {df_data['DataDate'].max()}\n")
        
        # Estadísticas de medicamentos por registro
        med_stats = df_data['medications_count'].describe()
        f.write(f"\nESTADÍSTICAS DE MEDICAMENTOS POR REGISTRO:\n")
        f.write(f"  Promedio: {med_stats['mean']:.2f} medicamentos\n")
        f.write(f"  Mediana: {med_stats['50%']:.0f} medicamentos\n")
        f.write(f"  Máximo: {med_stats['max']:.0f} medicamentos\n")
        f.write(f"  Mínimo: {med_stats['min']:.0f} medicamentos\n")
        
        # Distribución de cantidad de medicamentos
        f.write(f"\nDISTRIBUCIÓN DE CANTIDAD DE MEDICAMENTOS:\n")
        med_dist = df_data['medications_count'].value_counts().sort_index().head(10)
        for count, freq in med_dist.items():
            f.write(f"  {count} medicamento(s): {freq:,} registros\n")
        
        # Estadísticas de pacientes
        f.write(f"\nESTADÍSTICAS DE PACIENTES:\n")
        f.write(f"  Pacientes únicos: {df_data['PatientId'].nunique():,}\n")
        f.write(f"  Promedio prescripciones agrupadas por paciente: {len(df_data)/df_data['PatientId'].nunique():.2f}\n")
        
        f.write(f"\nFORMATO DE ARCHIVOS:\n")
        f.write(f"  Columna A: Import Clinic Record ID (correlativo desde 99883852)\n")
        f.write(f"  Columna B: Import Pet ID (PatientId original)\n")
        f.write(f"  Columna C: Import Date (fecha de prescripción)\n")
        f.write(f"  Columna D: Note (Medicamentos agrupados con formato específico)\n")
        f.write(f"\nFORMATO DE NOTA:\n")
        f.write(f"  Name\\nDescription\\nRequestedUsage + AmountToBuy[PARRAFO]...\n")
        f.write(f"  [PARRAFO] - Separador entre medicamentos del mismo registro\n")
    
    print(f"✅ Reporte guardado: {summary_file}")

def generate_validation_samples(df_data, output_dir):
    """Generar muestras para validación"""
    print(f"\n🔍 Generando muestras de validación...")
    
    samples_file = os.path.join(output_dir, "muestras_validacion.txt")
    
    with open(samples_file, 'w', encoding='utf-8') as f:
        f.write("MUESTRAS DE VALIDACIÓN - PRESCRIPCIONES\n")
        f.write("=" * 50 + "\n\n")
        
        # Muestras con múltiples medicamentos
        multi_meds = df_data[df_data['medications_count'] > 1].head(3)
        f.write("EJEMPLOS CON MÚLTIPLES MEDICAMENTOS:\n")
        f.write("-" * 35 + "\n")
        for i, (_, row) in enumerate(multi_meds.iterrows(), 1):
            f.write(f"Ejemplo {i} ({row['medications_count']} medicamentos):\n")
            f.write(f"  ID: {row['import_clinic_record_id']}\n")
            f.write(f"  Paciente: {row['PatientId']}\n")
            f.write(f"  Fecha: {row['DataDate']}\n")
            f.write(f"  Nota completa:\n")
            # Mostrar nota con saltos de línea para mejor legibilidad
            note_formatted = row['Note'].replace('[PARRAFO]', '\n--- MEDICAMENTO SEPARADO ---\n')
            f.write(f"    {note_formatted[:500]}{'...' if len(row['Note']) > 500 else ''}\n\n")
        
        # Muestras con un solo medicamento
        single_meds = df_data[df_data['medications_count'] == 1].head(3)
        f.write("EJEMPLOS CON UN SOLO MEDICAMENTO:\n")
        f.write("-" * 32 + "\n")
        for i, (_, row) in enumerate(single_meds.iterrows(), 1):
            f.write(f"Ejemplo {i}:\n")
            f.write(f"  ID: {row['import_clinic_record_id']}\n")
            f.write(f"  Paciente: {row['PatientId']}\n")
            f.write(f"  Fecha: {row['DataDate']}\n")
            f.write(f"  Nota: {row['Note'][:200]}{'...' if len(row['Note']) > 200 else ''}\n\n")
    
    print(f"✅ Muestras guardadas: {samples_file}")

def main():
    print("🏥 GENERADOR DE TEMPLATES - PRESCRIPCIONES")
    print("=" * 50)
    
    output_dir = "/Users/enrique/Proyectos/imports/generated_files/prescripciones"
    
    try:
        # 1. Cargar datos
        df_active = load_and_prepare_data()
        
        # 2. Agrupar por PatientId + DataDate
        grouped_data = group_prescriptions_by_datetime(df_active)
        
        # 3. Procesar registros agrupados
        df_result = process_prescription_records(grouped_data)
        
        if len(df_result) == 0:
            print("❌ No hay registros para procesar")
            return
        
        # 4. Crear templates Excel
        generated_files = create_excel_template(df_result, output_dir)
        
        # 5. Generar reporte resumen
        generate_summary_report(df_result, df_active, output_dir)
        
        # 6. Generar muestras de validación
        generate_validation_samples(df_result, output_dir)
        
        # 7. Resumen final
        print(f"\n🎉 ¡GENERACIÓN COMPLETADA!")
        print(f"📁 Directorio: {output_dir}")
        print(f"📄 Archivos generados: {len(generated_files)}")
        print(f"📊 Registros agrupados: {len(df_result):,}")
        print(f"💊 Prescripciones originales: {len(df_active):,}")
        print(f"📈 Factor de compresión: {len(df_active)/len(df_result):.2f}x")
        print(f"🏥 Pacientes únicos: {df_result['PatientId'].nunique():,}")
        print(f"🆔 Rango IDs: {df_result['import_clinic_record_id'].min()} - {df_result['import_clinic_record_id'].max()}")
        
        # Estadísticas de agrupación
        med_stats = df_result['medications_count']
        print(f"💊 Medicamentos por registro:")
        print(f"   Promedio: {med_stats.mean():.2f}")
        print(f"   Máximo: {med_stats.max()}")
        
        # Mostrar archivos generados
        print(f"\n📋 Archivos creados:")
        for filepath in generated_files:
            filename = os.path.basename(filepath)
            print(f"   • {filename}")
        print(f"   • resumen_prescripciones_import.txt")
        print(f"   • muestras_validacion.txt")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
