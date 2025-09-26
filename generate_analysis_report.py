#!/usr/bin/env python3
"""
Generador de reporte Excel con análisis completo de cuvet-v2.xlsx
Incluye comparación con versión original, rangos de fechas, estadísticas de eliminación, etc.
"""

import pandas as pd
import numpy as np
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_analysis_report():
    print("📊 GENERANDO REPORTE DE ANÁLISIS EN EXCEL")
    print("=" * 60)
    
    # Archivos a analizar
    original_file = "/Users/enrique/Proyectos/imports/source/cuvet.xlsx"
    new_file = "/Users/enrique/Proyectos/imports/source/cuvet-v2.xlsx"
    output_file = "/Users/enrique/Proyectos/imports/analisis_cuvet_completo.xlsx"
    
    # Verificar archivos
    if not os.path.exists(new_file):
        print(f"❌ Archivo no encontrado: {new_file}")
        return
    
    # Crear workbook
    wb = Workbook()
    
    # Datos para el análisis
    analysis_data = {
        'resumen_general': [],
        'analisis_por_pestana': [],
        'estadisticas_eliminacion': [],
        'rangos_fechas': [],
        'nuevas_pestanas': []
    }
    
    print("📖 Analizando archivo cuvet-v2.xlsx...")
    
    try:
        # Leer información del archivo nuevo
        new_sheets = pd.ExcelFile(new_file).sheet_names
        
        # Si existe el archivo original, comparar
        has_original = os.path.exists(original_file)
        if has_original:
            original_sheets = pd.ExcelFile(original_file).sheet_names
            common_sheets = [sheet for sheet in original_sheets if sheet in new_sheets]
            new_only_sheets = [sheet for sheet in new_sheets if sheet not in original_sheets]
        else:
            original_sheets = []
            common_sheets = []
            new_only_sheets = new_sheets
        
        # Analizar cada pestaña
        total_records = 0
        total_deleted = 0
        total_active = 0
        
        for sheet_name in new_sheets:
            print(f"   📋 Procesando: {sheet_name}")
            
            try:
                df = pd.read_excel(new_file, sheet_name=sheet_name, engine='openpyxl')
                sheet_records = len(df)
                total_records += sheet_records
                
                # Análisis básico
                sheet_analysis = {
                    'Pestaña': sheet_name,
                    'Registros': sheet_records,
                    'Estado': 'Nueva' if sheet_name in new_only_sheets else 'Existente',
                    'Columnas': len(df.columns),
                    'Tiene_IsDeleted': 'IsDeleted' in df.columns
                }
                
                # Análisis de eliminación
                if 'IsDeleted' in df.columns:
                    deleted_count = df['IsDeleted'].sum()
                    active_count = sheet_records - deleted_count
                    deletion_rate = (deleted_count / sheet_records) * 100 if sheet_records > 0 else 0
                    
                    total_deleted += deleted_count
                    total_active += active_count
                    
                    sheet_analysis.update({
                        'Eliminados': deleted_count,
                        'Activos': active_count,
                        'Tasa_Eliminacion': f"{deletion_rate:.1f}%"
                    })
                    
                    # Agregar a estadísticas de eliminación
                    analysis_data['estadisticas_eliminacion'].append({
                        'Pestaña': sheet_name,
                        'Total': sheet_records,
                        'Activos': active_count,
                        'Eliminados': deleted_count,
                        'Porcentaje_Eliminacion': deletion_rate,
                        'Estado': 'Nueva' if sheet_name in new_only_sheets else 'Existente'
                    })
                else:
                    sheet_analysis.update({
                        'Eliminados': 'N/A',
                        'Activos': sheet_records,
                        'Tasa_Eliminacion': 'N/A'
                    })
                
                # Análisis de fechas
                date_columns = []
                for col in df.columns:
                    if any(keyword in col.lower() for keyword in ['date', 'fecha', 'modified', 'created']):
                        date_columns.append(col)
                
                if date_columns:
                    for date_col in date_columns:
                        try:
                            df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
                            valid_dates = df[date_col].dropna()
                            
                            if len(valid_dates) > 0:
                                analysis_data['rangos_fechas'].append({
                                    'Pestaña': sheet_name,
                                    'Columna_Fecha': date_col,
                                    'Fecha_Minima': valid_dates.min(),
                                    'Fecha_Maxima': valid_dates.max(),
                                    'Registros_Validos': len(valid_dates),
                                    'Registros_Nulos': len(df) - len(valid_dates),
                                    'Estado': 'Nueva' if sheet_name in new_only_sheets else 'Existente'
                                })
                        except:
                            continue
                
                # Información de nuevas pestañas
                if sheet_name in new_only_sheets:
                    # Obtener tipos de datos más comunes
                    sample_data = []
                    for col in df.columns[:5]:  # Primeras 5 columnas
                        non_null_values = df[col].dropna()
                        if len(non_null_values) > 0:
                            sample_data.append(f"{col}: {non_null_values.iloc[0]}")
                    
                    analysis_data['nuevas_pestanas'].append({
                        'Pestaña': sheet_name,
                        'Registros': sheet_records,
                        'Columnas': len(df.columns),
                        'Principales_Columnas': ', '.join(df.columns[:5].tolist()),
                        'Datos_Muestra': ' | '.join(sample_data[:3]),
                        'Periodo_Datos': f"Desde {df[date_columns[0]].min()}" if date_columns else 'N/A'
                    })
                
                analysis_data['analisis_por_pestana'].append(sheet_analysis)
                
            except Exception as e:
                print(f"❌ Error procesando {sheet_name}: {e}")
                continue
        
        # Resumen general
        analysis_data['resumen_general'] = [{
            'Metrica': 'Total de Pestañas',
            'Valor': len(new_sheets),
            'Detalle': f"{len(common_sheets)} existentes + {len(new_only_sheets)} nuevas"
        }, {
            'Metrica': 'Total de Registros',
            'Valor': total_records,
            'Detalle': f"{total_records:,} registros en total"
        }, {
            'Metrica': 'Registros Activos',
            'Valor': total_active,
            'Detalle': f"{total_active:,} registros activos"
        }, {
            'Metrica': 'Registros Eliminados',
            'Valor': total_deleted,
            'Detalle': f"{total_deleted:,} registros eliminados"
        }, {
            'Metrica': 'Tasa de Eliminación General',
            'Valor': f"{(total_deleted/total_records)*100:.1f}%" if total_records > 0 else "0%",
            'Detalle': f"Promedio de eliminación en todas las pestañas"
        }, {
            'Metrica': 'Nuevas Pestañas',
            'Valor': len(new_only_sheets),
            'Detalle': ', '.join(new_only_sheets) if new_only_sheets else 'Ninguna'
        }, {
            'Metrica': 'Fecha de Análisis',
            'Valor': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'Detalle': 'Momento de generación del reporte'
        }]
        
        # Crear las hojas del Excel
        create_excel_sheets(wb, analysis_data)
        
        # Guardar archivo
        wb.save(output_file)
        print(f"✅ Reporte generado exitosamente: {output_file}")
        
        return output_file
        
    except Exception as e:
        print(f"❌ Error general: {e}")
        return None

def create_excel_sheets(wb, data):
    """Crear las hojas del Excel con formato"""
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Resumen General
    ws1 = wb.create_sheet("📊 Resumen General")
    df_resumen = pd.DataFrame(data['resumen_general'])
    
    # Headers
    headers = ['Métrica', 'Valor', 'Detalle']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Data
    for row_idx, row_data in enumerate(data['resumen_general'], 2):
        ws1.cell(row=row_idx, column=1, value=row_data['Metrica']).border = border
        ws1.cell(row=row_idx, column=2, value=str(row_data['Valor'])).border = border
        ws1.cell(row=row_idx, column=3, value=row_data['Detalle']).border = border
    
    # Ajustar columnas
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 50
    
    # 2. Análisis por Pestaña
    ws2 = wb.create_sheet("📋 Análisis por Pestaña")
    df_pestanas = pd.DataFrame(data['analisis_por_pestana'])
    
    if not df_pestanas.empty:
        # Headers
        for col, header in enumerate(df_pestanas.columns, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Data
        for row_idx, (_, row_data) in enumerate(df_pestanas.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = border
                if row_data['Estado'] == 'Nueva':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Ajustar columnas
        for col in ws2.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws2.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
    
    # 3. Estadísticas de Eliminación
    ws3 = wb.create_sheet("🗑️ Eliminaciones")
    df_eliminacion = pd.DataFrame(data['estadisticas_eliminacion'])
    
    if not df_eliminacion.empty:
        # Headers
        for col, header in enumerate(df_eliminacion.columns, 1):
            cell = ws3.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Data
        for row_idx, (_, row_data) in enumerate(df_eliminacion.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                
                # Colorear según tasa de eliminación
                if col_idx == 5 and isinstance(value, (int, float)):  # Porcentaje_Eliminacion
                    if value > 5:
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif value > 3:
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                
                if row_data['Estado'] == 'Nueva':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Ajustar columnas
        for col in ws3.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws3.column_dimensions[col[0].column_letter].width = min(max_length + 2, 20)
    
    # 4. Rangos de Fechas
    ws4 = wb.create_sheet("📅 Rangos de Fechas")
    df_fechas = pd.DataFrame(data['rangos_fechas'])
    
    if not df_fechas.empty:
        # Headers
        for col, header in enumerate(df_fechas.columns, 1):
            cell = ws4.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border
        
        # Data
        for row_idx, (_, row_data) in enumerate(df_fechas.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                if isinstance(value, pd.Timestamp):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws4.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = border
                if row_data['Estado'] == 'Nueva':
                    cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Ajustar columnas
        for col in ws4.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws4.column_dimensions[col[0].column_letter].width = min(max_length + 2, 25)
    
    # 5. Nuevas Pestañas
    ws5 = wb.create_sheet("🆕 Nuevas Pestañas")
    df_nuevas = pd.DataFrame(data['nuevas_pestanas'])
    
    if not df_nuevas.empty:
        # Headers
        for col, header in enumerate(df_nuevas.columns, 1):
            cell = ws5.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = border
        
        # Data
        for row_idx, (_, row_data) in enumerate(df_nuevas.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws5.cell(row=row_idx, column=col_idx, value=str(value))
                cell.border = border
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
        
        # Ajustar columnas
        for col in ws5.columns:
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            ws5.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

if __name__ == "__main__":
    output_file = create_analysis_report()
    if output_file:
        print(f"\n🎉 ¡Reporte Excel generado exitosamente!")
        print(f"📁 Ubicación: {output_file}")
        print(f"📊 El archivo contiene 5 hojas con análisis detallado")
    else:
        print("❌ Error al generar el reporte")
